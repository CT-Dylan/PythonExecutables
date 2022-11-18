from os import getcwd,listdir,devnull
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
from xlrd import open_workbook
from re import sub


# Get the values to change from the Excel File
path = getcwd()
tableName = path+"\\renameKeithleySheetsTable.xlsx"
t = load_workbook(tableName, data_only = True)
worksheet = t.active
replaceThis = []
byThat = []

for i in range(3, worksheet.max_row):
    if( worksheet[i][0].value is not None):
        replaceThis.insert(i-3, worksheet[i][0].value)
        byThat.insert(i-3, worksheet[i][1].value)


#Create new Excel file if file does not exist
didExist = 1
if "CHIP_PleaseRenameFile.xlsx" not in listdir():
    didExist = 0
    newWorkbook = Workbook()
    newWorkbook.save(path+"\\CHIP_PleaseRenameFile.xlsx")
    newWorkbook.close()

#Open new Excel file
dst_wb = load_workbook(path+"\\CHIP_PleaseRenameFile.xlsx")
 
# Get the files to change
for count, f in enumerate(listdir()):
    if '.xls' in f and f is not tableName and f != "renameKeithleySheetsTable.xlsx" and f != "CHIP_PleaseRenameFile.xlsx":
        print(str(f))
        src_wb = open_workbook(f, logfile=open(devnull, 'w'), formatting_info=True)
        for s in src_wb.sheet_names():
            src_ws = src_wb.sheet_by_name(s)
            eleNum = 0
            for ele in replaceThis:
                if ele == s:
                    dst_ws = dst_wb.create_sheet(byThat[eleNum])
                    for row in range(0,src_ws.nrows):
                        for column in range(0,src_ws.ncols):
                            if isinstance(src_ws.cell_value(row,column),str):
                                dst_ws[get_column_letter(column+1) + str(row+1)].value = sub('^=', ' =', src_ws.cell_value(row,column))
                            else:
                                dst_ws[get_column_letter(column+1) + str(row+1)].value = src_ws.cell_value(row,column)
                eleNum += 1
        src_wb.release_resources()
                
if didExist == 0:
    dst_wb.remove(dst_wb["Sheet"])         
 
if(len(dst_wb.sheetnames) > 0): 
    dst_wb.save(path+"\\CHIP_PleaseRenameFile.xlsx")
dst_wb.close()
    
t.close()

 